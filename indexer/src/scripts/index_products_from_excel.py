from __future__ import annotations

"""Script de indexación de productos desde Excel a Azure AI Search.

Lee un archivo Excel con columnas en español ["nombre", "precio_venta",
"descripcion", "tienda", "imagenes"], prepara documentos con el servicio
`AzureProductSearchService` y los sube al índice configurado.

Uso:
    python index_products_from_excel.py --file indexer/data/productos_ejemplo.xlsx --recreate

Requisitos:
    - Variables de entorno para Azure Search (ver servicio)
    - `openpyxl` para leer Excel (`pip install openpyxl`)
"""

import argparse
import asyncio
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    load_workbook = None  # type: ignore

# Asegurar que el proyecto raíz esté en sys.path para que `mcp` sea importable
PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# Importar el servicio del indexador
from indexer.src.services.ai_search_service import AzureProductSearchService  # type: ignore


@dataclass
class ExcelReadConfig:
    """Configuración para lectura del Excel.

    Permite definir el nombre de las columnas esperadas y si son obligatorias.
    """

    name_col: str = "nombre"
    price_col: str = "precio_venta"
    description_col: str = "descripcion"
    store_col: str = "tienda"
    images_col: str = "imagenes"


def parse_args() -> argparse.Namespace:
    """Parsea argumentos de línea de comandos.

    Returns:
        argparse.Namespace: Parámetros del usuario (ruta de archivo, recreate, limit).
    """
    parser = argparse.ArgumentParser(description="Indexar productos desde Excel a Azure AI Search")
    default_excel = PROJECT_ROOT / "indexer" / "data" / "productos_ejemplo.xlsx"
    parser.add_argument(
        "--file",
        type=Path,
        default=default_excel,
        help=f"Ruta del Excel de productos (por defecto: {default_excel})",
    )
    parser.add_argument(
        "--recreate",
        action="store_true",
        help="Eliminar y recrear el índice antes de indexar",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limitar el número de filas a procesar (debug)",
    )
    return parser.parse_args()


def read_products_from_excel(file_path: Path, config: Optional[ExcelReadConfig] = None) -> List[Dict[str, Any]]:
    """Lee productos desde un archivo Excel y retorna filas como diccionarios.

    Args:
        file_path: Ruta del archivo Excel.
        config: Configuración de columnas esperadas.

    Returns:
        list[dict]: Filas con llaves en español según el Excel.
    """
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl no está instalado. Instala con: pip install openpyxl"
        )

    cfg = config or ExcelReadConfig()

    wb = load_workbook(filename=str(file_path), data_only=True)
    ws = wb.active

    # Mapear encabezados -> índice de columna
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(h).strip().lower(): idx for idx, h in enumerate(header_row) if h is not None}

    required = [cfg.name_col, cfg.price_col, cfg.description_col, cfg.store_col, cfg.images_col]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(f"Faltan columnas requeridas en el Excel: {missing}")

    products: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Saltar filas completamente vacías
        if row is None or all(cell in (None, "") for cell in row):
            continue

        def get(col_name: str) -> Any:
            idx = headers.get(col_name)
            return row[idx] if idx is not None and idx < len(row) else None

        nombre = (get(cfg.name_col) or "").strip() if isinstance(get(cfg.name_col), str) else get(cfg.name_col)
        precio = get(cfg.price_col)
        descripcion = (get(cfg.description_col) or "").strip() if isinstance(get(cfg.description_col), str) else get(cfg.description_col)
        tienda = (get(cfg.store_col) or "").strip() if isinstance(get(cfg.store_col), str) else get(cfg.store_col)
        imagenes = get(cfg.images_col) or ""

        # Normalizar precio a float cuando sea posible
        price_value: Optional[float]
        try:
            price_value = float(precio) if precio not in (None, "") else None
        except Exception:
            price_value = None

        products.append(
            {
                "nombre": nombre or "",
                "precio_venta": price_value,
                "descripcion": descripcion or "",
                "tienda": tienda or "",
                "imagenes": str(imagenes) if imagenes is not None else "",
            }
        )

    return products


async def run(file_path: Path, recreate: bool = False, limit: Optional[int] = None) -> int:
    """Ejecuta el flujo de indexación de productos.

    Args:
        file_path: Ruta al Excel de productos.
        recreate: Si es True, recrea el índice antes de indexar.
        limit: Límite opcional de filas a procesar.

    Returns:
        int: Código de salida (0=ok, !=0 error).
    """
    if not file_path.exists():
        print(f"[ERROR] No se encontró el archivo: {file_path}")
        return 2

    try:
        rows = read_products_from_excel(file_path)
    except Exception as exc:
        print(f"[ERROR] Falló la lectura del Excel: {exc}")
        return 3

    if limit is not None:
        rows = rows[: max(0, limit)]

    if not rows:
        print("[INFO] No hay filas para indexar")
        return 0

    service = AzureProductSearchService()

    if recreate:
        result = await service.create_index(force_recreate=True)
    else:
        result = await service.create_index(force_recreate=False)

    if not result.get("success"):
        print(f"[ERROR] No se pudo preparar el índice: {result.get('error')}")
        return 4

    prepared: List[Dict[str, Any]] = []
    for r in rows:
        doc = await service.prepare_product_document(r)
        prepared.append(doc)

    upload_result = await service.upload_documents(prepared)
    if not upload_result.get("success"):
        print(f"[ERROR] Falló la subida de documentos: {upload_result.get('error')}")
        return 5

    print(
        f"[OK] Indexación completada. Documentos subidos: {upload_result.get('uploaded', 0)}"
    )
    return 0


def main() -> None:
    """Punto de entrada del script CLI."""
    args = parse_args()
    exit_code = asyncio.run(run(args.file, recreate=args.recreate, limit=args.limit))
    raise SystemExit(exit_code)


if __name__ == "__main__":
    main()


# Correr con:
#python indexer/src/scripts/index_products_from_excel.py --file indexer/data/productos_ejemplo.xlsx --recreate